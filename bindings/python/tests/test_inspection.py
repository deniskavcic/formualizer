from __future__ import annotations

from collections.abc import Mapping

import pytest

import formualizer as fz


def build_workbook() -> fz.Workbook:
    wb = fz.Workbook()
    sheet = wb.sheet("Sheet1")
    sheet.set_value(1, 1, 10)
    sheet.set_value(2, 1, 20)
    sheet.set_value(2, 2, 5)
    sheet.set_formula(1, 2, "=A2+A1+SUM(A1:A2)")
    sheet.set_formula(1, 3, "=A1*2")
    sheet.set_formula(1, 4, "=A1*3")
    sheet.set_formula(1, 5, "=A1*4")
    sheet.set_formula(1, 6, "=G1+1")
    sheet.set_formula(1, 7, "=F1+1")
    sheet.set_formula(1, 8, "=SEQUENCE(3)")
    wb.evaluate_all()
    return wb


def test_all_inspection_entry_points_and_owned_reports() -> None:
    wb = build_workbook()

    snapshot = wb.inspect_cell("Sheet1!B1")
    assert snapshot.cell.address == "Sheet1!B1"
    assert snapshot.cell.value_included is True
    assert snapshot.cell.staleness is fz.Staleness.Current
    assert isinstance(snapshot.stamp.mutation_revision, int)
    assert isinstance(snapshot.stamp.recalc_epoch, int)
    assert snapshot.to_dict()["cell"]["address"] == "Sheet1!B1"

    precedents = wb.precedents("Sheet1!B1")
    assert [item.reference.kind for item in precedents.precedents] == [
        fz.ReferenceKind.Cell,
        fz.ReferenceKind.Cell,
        fz.ReferenceKind.Range,
    ]
    assert [item.reference.address for item in precedents.precedents[:2]] == [
        "Sheet1!A2",
        "Sheet1!A1",
    ]
    assert precedents.precedents[2].reference.declared == "Sheet1!A1:A2"
    assert precedents.precedents[2].reference.resolved == "Sheet1!A1:A2"
    assert precedents.precedents[2].reference.cell_count == 2
    assert all(
        item.provenance is fz.Provenance.Declared for item in precedents.precedents
    )

    page = wb.range_page("Sheet1!A1:B2", limit=3)
    assert page.declared == "Sheet1!A1:B2"
    assert page.resolved == "Sheet1!A1:B2"
    assert page.total == 4
    assert [item.address for item in page.items] == [
        "Sheet1!A1",
        "Sheet1!B1",
        "Sheet1!A2",
    ]
    assert page.next_offset == 3
    final_page = wb.range_page(
        "Sheet1!A1:B2", offset=page.next_offset, limit=3, expected_stamp=page.stamp
    )
    assert [item.address for item in final_page.items] == ["Sheet1!B2"]
    assert final_page.next_offset is None

    trace = wb.trace(["Sheet1!B1"], max_depth=2, max_nodes=20)
    assert trace.roots[0].address == "Sheet1!B1"
    assert trace.nodes["Sheet1!A1"].address == "Sheet1!A1"
    assert list(trace.nodes)[0] == "Sheet1!B1"
    assert len(trace.nodes) == len(trace.to_dict()["nodes"])
    assert trace.links

    old = snapshot
    wb.set_formula("Sheet1", 1, 2, "=A1+1")
    dirty = wb.inspect_cell("Sheet1!B1")
    assert dirty.cell.staleness is fz.Staleness.Dirty
    assert old.cell.staleness is fz.Staleness.Current
    assert old.cell.formula != dirty.cell.formula


def test_all_option_kwargs_are_plumbed_and_zero_budgets_are_in_band() -> None:
    wb = fz.Workbook()
    sheet = wb.sheet("S")
    for row in range(1, 6):
        sheet.set_value(row, 1, row)
    sheet.set_formula(1, 2, "=SUM(A1:A4)")
    sheet.set_formula(1, 3, "=B1+1")
    sheet.set_formula(1, 4, "=C1+1")
    sheet.set_formula(1, 5, "=D1+1")
    wb.evaluate_all()

    assert len(wb.dependents("S!A1", max_results=1).dependents) == 1
    assert len(wb.precedents("S!B1", max_links=0).precedents) == 0
    assert wb.dependents("S!A1", max_work=0).truncation.incomplete is True

    shallow = wb.trace(["S!E1"], max_depth=1)
    deep = wb.trace(["S!E1"], max_depth=4)
    assert len(shallow.nodes) < len(deep.nodes)
    assert wb.trace(["S!E1"], max_nodes=1, max_depth=0).roots[0].address == "S!E1"
    assert len(wb.trace(["S!B1"], max_links=0).links) == 0
    assert wb.trace(["S!B1"], max_work=0).truncation.incomplete is True

    ranged = wb.trace(["S!B1"], max_depth=1, range_member_budget=1)
    ranged_link = next(
        link for link in ranged.links if link.reference.kind is fz.ReferenceKind.Range
    )
    assert len(ranged_link.targets) == 1
    assert ranged_link.omitted.kind is fz.OmittedCountKind.Exact
    assert ranged_link.omitted.count == 3

    for report_cells in (
        [wb.inspect_cell("S!B1", include_values=False).cell],
        [node.cell for node in wb.trace(["S!B1"], include_values=False).nodes.values()],
        wb.range_page("S!A1:B1", include_values=False).items,
    ):
        assert report_cells
        assert all(
            not cell.value_included and cell.value is None for cell in report_cells
        )

    page = wb.range_page("S!A1:B1", offset=1, limit=1)
    assert page.offset == 1
    assert len(page.items) == 1
    wb.set_value("S", 6, 1, 6)
    with pytest.raises(fz.InspectionRevisionMismatchError):
        wb.range_page("S!A1:B1", expected_stamp=page.stamp)


def test_trace_direction_depth_root_order_and_duplicate_semantics() -> None:
    wb = fz.Workbook()
    sheet = wb.sheet("S")
    sheet.set_value(1, 1, 1)
    sheet.set_formula(1, 2, "=A1+1")
    sheet.set_formula(1, 3, "=B1+1")
    sheet.set_formula(1, 4, "=C1+1")
    wb.evaluate_all()

    precedents = wb.trace(["S!B1"], direction=fz.TraceDirection.Precedents)
    dependents = wb.trace(["S!B1"], direction=fz.TraceDirection.Dependents)
    assert precedents.direction is fz.TraceDirection.Precedents
    assert dependents.direction is fz.TraceDirection.Dependents
    assert set(precedents.nodes) == {"S!B1", "S!A1"}
    assert set(dependents.nodes) == {"S!B1", "S!C1", "S!D1"}

    ordered = wb.trace(["S!D1", "S!B1", "S!C1"])
    assert [root.address for root in ordered.roots] == ["S!D1", "S!B1", "S!C1"]
    duplicated = wb.trace(["S!B1", "S!B1", "S!C1"])
    unique = wb.trace(["S!B1", "S!C1"])
    assert [root.address for root in duplicated.roots] == ["S!B1", "S!B1", "S!C1"]
    assert len(duplicated.nodes) == len(unique.nodes)


def test_dependents_truncation_retains_address_least_across_sheets() -> None:
    wb = fz.Workbook()
    zzz = wb.sheet("ZZZ")
    aaa = wb.sheet("AAA")
    mid = wb.sheet("Mid")
    zzz.set_value(1, 1, 1)
    for sheet, column in ((zzz, 2), (aaa, 1), (mid, 1)):
        sheet.set_formula(1, column, "=ZZZ!A1*2")
    wb.evaluate_all()

    full = wb.dependents("ZZZ!A1", max_results=100)
    top = wb.dependents("ZZZ!A1", max_results=1)
    assert [item.cell for item in full.dependents] == ["AAA!A1", "Mid!A1", "ZZZ!B1"]
    assert [item.cell for item in top.dependents] == ["AAA!A1"]
    assert top.truncation.incomplete is True


def test_trace_nodes_is_normalized_mapping_with_multisheet_keys() -> None:
    wb = fz.Workbook()
    source = wb.sheet("My Sheet")
    target = wb.sheet("Sheet2")
    target.set_value(1, 1, 7)
    source.set_formula(1, 2, "=Sheet2!A1+1")
    wb.evaluate_all()

    nodes = wb.trace(["My Sheet!B1"]).nodes
    assert isinstance(nodes, Mapping)
    assert nodes["My Sheet!B1"].address == "'My Sheet'!B1"
    assert "My Sheet!B1" in nodes
    assert nodes.get("My Sheet!B1").address == "'My Sheet'!B1"
    assert nodes.get("My Sheet!Z99") is None
    assert set(nodes.keys()) == {"'My Sheet'!B1", "Sheet2!A1"}
    assert {node.address for node in nodes.values()} == set(nodes.keys())
    assert {key for key, _ in nodes.items()} == set(nodes.keys())
    assert nodes["Sheet2!A1"].address == "Sheet2!A1"
    with pytest.raises(KeyError):
        nodes["'My Sheet'!A1"]


def test_stamp_value_semantics_and_revision_fields() -> None:
    wb = build_workbook()
    report = wb.inspect_cell("Sheet1!B1")
    same_call_left = report.stamp
    same_call_right = report.stamp
    cross_call = wb.range_page("Sheet1!A1:B1").stamp

    assert same_call_left == same_call_right == cross_call
    assert same_call_left.__ne__(cross_call) is False
    assert {same_call_left: "fresh"}[cross_call] == "fresh"
    assert hash(same_call_left) == hash(cross_call)
    assert report.stamp == wb.inspect_cell("Sheet1!B1").stamp
    wb.range_page("Sheet1!A1:B1", expected_stamp=report.stamp)

    before = report.stamp
    wb.set_value("Sheet1", 9, 9, 99)
    after = wb.inspect_cell("Sheet1!B1").stamp
    assert after.mutation_revision != before.mutation_revision
    assert after.recalc_epoch == before.recalc_epoch
    assert after != before
    with pytest.raises(TypeError):
        hash(report.cell)


def _cell_dict(address: str, formula: str | None, value: float) -> dict[str, object]:
    return {
        "address": address,
        "formula": formula,
        "value": value,
        "value_included": True,
        "staleness": "Current",
        "volatile": False,
        "spill": None,
    }


def _cell_reference(address: str) -> dict[str, object]:
    return {
        "kind": "Cell",
        "address": address,
        "declared": None,
        "resolved": None,
        "cell_count": None,
        "name": None,
        "specifier": None,
        "raw": None,
        "text": None,
        "reason": None,
        "resolution": None,
    }


def test_to_dict_full_nested_structure_for_every_report_kind() -> None:
    wb = fz.Workbook()
    sheet = wb.sheet("S")
    sheet.set_value(1, 1, 1)
    sheet.set_value(2, 1, 2)
    sheet.set_formula(1, 2, "=A1+SUM(A1:A2)")
    sheet.set_formula(1, 3, "=B1+1")
    wb.evaluate_all()
    stamp = {"mutation_revision": 6, "recalc_epoch": 1}

    assert wb.inspect_cell("S!B1").to_dict() == {
        "stamp": stamp,
        "cell": _cell_dict("S!B1", "=A1 + SUM(A1:A2)", 4.0),
    }
    assert wb.precedents("S!B1", max_links=1).to_dict() == {
        "stamp": stamp,
        "cell": "S!B1",
        "precedents": [
            {"reference": _cell_reference("S!A1"), "provenance": "Declared"}
        ],
        "truncation": {
            "incomplete": True,
            "omitted": {"kind": "AtLeast", "count": 1},
        },
    }
    assert wb.dependents("S!A1", max_results=1).to_dict() == {
        "stamp": stamp,
        "cell": "S!A1",
        "dependents": [{"cell": "S!B1", "via": []}],
        "truncation": {"incomplete": False, "omitted": None},
    }

    link = {
        "source": "S!B1",
        "reference": _cell_reference("S!A1"),
        "kind": {"kind": "Formula", "provenance": "Declared"},
        "targets": [{"node": "S!A1", "disposition": "Expanded"}],
        "omitted": None,
    }
    assert wb.trace(
        ["S!B1"], max_depth=1, max_links=1, range_member_budget=1
    ).to_dict() == {
        "stamp": stamp,
        "direction": "Precedents",
        "roots": ["S!B1"],
        "nodes": {
            "S!B1": {
                "id": 0,
                "cell": _cell_dict("S!B1", "=A1 + SUM(A1:A2)", 4.0),
                "links": [link],
            },
            "S!A1": {"id": 1, "cell": _cell_dict("S!A1", None, 1.0), "links": []},
        },
        "links": [link],
        "truncation": {
            "incomplete": True,
            "omitted": {"kind": "AtLeast", "count": 1},
        },
    }
    assert wb.range_page("S!A1:B1", limit=1).to_dict() == {
        "stamp": stamp,
        "declared": "S!A1:B1",
        "resolved": "S!A1:B1",
        "total": 2,
        "offset": 0,
        "items": [_cell_dict("S!A1", None, 1.0)],
        "next_offset": 1,
    }


def test_spill_payload_provenance_omission_cell_count_and_volatile() -> None:
    wb = fz.Workbook()
    sheet = wb.sheet("S")
    for row in range(1, 5):
        sheet.set_value(row, 1, row)
    sheet.set_formula(1, 2, "=SUM(A1:A4)")
    sheet.set_formula(1, 8, "=SEQUENCE(3)")
    sheet.set_formula(2, 9, "=H2*2")
    sheet.set_formula(1, 10, "=RAND()")
    wb.evaluate_all()

    assert wb.inspect_cell("S!J1").cell.volatile is True
    assert wb.inspect_cell("S!B1").cell.volatile is False
    precedent = wb.precedents("S!B1").precedents[0]
    assert precedent.provenance is fz.Provenance.Declared
    assert precedent.reference.cell_count == 4

    dependent = wb.dependents("S!H1").dependents[0]
    assert dependent.cell == "S!I2"
    assert dependent.via == ["S!H2"]

    spill_link = wb.trace(["S!H2"], max_depth=2).links[0]
    assert spill_link.kind.kind is fz.TraceLinkKindType.SpillAnchor
    assert spill_link.kind.provenance is None
    assert spill_link.targets[0].node.address == "S!H1"
    assert spill_link.targets[0].disposition is fz.LinkDisposition.Expanded

    formula_link = wb.trace(["S!B1"], range_member_budget=1).links[0]
    assert formula_link.kind.kind is fz.TraceLinkKindType.Formula
    assert formula_link.kind.provenance is fz.Provenance.Declared
    assert formula_link.omitted.kind is fz.OmittedCountKind.Exact
    assert formula_link.omitted.count == 3


def test_range_text_quotes_every_consumer_surface(tmp_path) -> None:
    import openpyxl
    from openpyxl.workbook.defined_name import DefinedName

    for sheet_name, canonical_sheet in (
        ("My Sheet", "'My Sheet'"),
        ("2024", "'2024'"),
        ("It's", "'It''s'"),
        ("Has!Bang", "'Has!Bang'"),
    ):
        path = tmp_path / f"quoted-{len(sheet_name)}-{ord(sheet_name[0])}.xlsx"
        source = openpyxl.Workbook()
        sheet = source.active
        sheet.title = sheet_name
        sheet["A1"] = 1
        sheet["A2"] = 2
        sheet["B1"] = "=SUM(A1:A2)+SUM(Block)"
        sheet["C1"] = "=SEQUENCE(2)"
        quoted_name = "'" + sheet_name.replace("'", "''") + "'"
        source.defined_names.add(
            DefinedName("Block", attr_text=f"{quoted_name}!$A$1:$A$2")
        )
        source.save(path)

        wb = fz.Workbook.load_path(str(path))
        wb.evaluate_all()
        expected = f"{canonical_sheet}!A1:A2"
        root = f"{canonical_sheet}!B1"
        precedents = wb.precedents(root).precedents
        range_reference = next(
            item.reference
            for item in precedents
            if item.reference.kind is fz.ReferenceKind.Range
        )
        name_reference = next(
            item.reference
            for item in precedents
            if item.reference.kind is fz.ReferenceKind.Name
        )
        page = wb.range_page(expected)
        spill = wb.inspect_cell(f"{canonical_sheet}!C1").cell.spill

        assert page.resolved == expected
        assert page.to_dict()["resolved"] == expected
        assert range_reference.resolved == expected
        assert range_reference.to_dict()["resolved"] == expected
        assert name_reference.resolution.resolved == expected
        assert name_reference.to_dict()["resolution"]["resolved"] == expected
        assert spill.extent == f"{canonical_sheet}!C1:C2"
        assert spill.to_dict()["extent"] == f"{canonical_sheet}!C1:C2"
        if sheet_name == "My Sheet":
            assert wb.inspect_cell(f"{sheet_name}!B1").cell.address == root


def test_malformed_addresses_use_inspection_error_hierarchy() -> None:
    wb = fz.Workbook()
    wb.sheet("S")
    malformed_cells = (
        "A1",
        "S!",
        "!A1",
        "S!ZZZZZZ1",
        "S!A0",
        "S!A99999999999999",
        "S!A1:B2",
        "S!$$",
        "S!1A",
    )
    for address in malformed_cells:
        with pytest.raises(fz.InspectionError) as caught:
            wb.inspect_cell(address)
        assert isinstance(caught.value, fz.InvalidInspectionAddressError)
        assert caught.value.code == "invalid_address"

    with pytest.raises(fz.InspectionError) as caught:
        wb.range_page("S!A1:B2:C3")
    assert isinstance(caught.value, fz.InvalidInspectionAddressError)
    assert caught.value.code == "invalid_address"

    with pytest.raises(
        fz.InvalidInspectionAddressError, match="unterminated sheet quote"
    ) as caught:
        wb.inspect_cell("'It's'!B1")
    assert caught.value.code == "invalid_address"
    with pytest.raises(fz.InvalidInspectionAddressError):
        wb.range_page("S!A1:1")


def test_revision_mismatch_has_structured_stamps() -> None:
    wb = build_workbook()
    page = wb.range_page("Sheet1!A1:B1")
    wb.set_value("Sheet1", 9, 9, 1)
    with pytest.raises(fz.InspectionRevisionMismatchError) as caught:
        wb.range_page("Sheet1!A1:B1", expected_stamp=page.stamp)
    assert caught.value.code == "revision_mismatch"
    assert caught.value.expected == page.stamp
    assert caught.value.actual == wb.inspect_cell("Sheet1!A1").stamp


def test_cycle_spill_mapping_repr_and_exception() -> None:
    wb = build_workbook()
    cycle = wb.trace(["Sheet1!F1"], max_depth=5, max_nodes=20)
    dispositions = [
        target.disposition for link in cycle.links for target in link.targets
    ]
    assert fz.LinkDisposition.Cycle in dispositions

    anchor = wb.inspect_cell("Sheet1!H1")
    member = wb.inspect_cell("Sheet1!H2")
    assert anchor.cell.spill.kind is fz.SpillRoleKind.Anchor
    assert anchor.cell.spill.extent == "Sheet1!H1:H3"
    assert member.cell.spill.kind is fz.SpillRoleKind.Member
    assert member.cell.spill.anchor == "Sheet1!H1"

    large_sheet = wb.sheet("Large")
    for column in range(1, 81):
        large_sheet.set_value(2, column, column)
    large_sheet.set_formula(1, 1, "=SUM(A2:CB2)")
    wb.evaluate_all()
    large = wb.trace(
        ["Large!A1"],
        max_depth=2,
        max_nodes=100,
        max_links=100,
        range_member_budget=90,
    )
    assert len(repr(large)) < 160
    assert "nodes=" in repr(large)
    assert "=A2" not in repr(large)

    with pytest.raises(fz.SheetNotFoundError) as caught:
        wb.inspect_cell("Missing!A1")
    assert caught.value.code == "sheet_not_found"
    assert caught.value.sheet == "Missing"


def test_defined_name_is_reported_from_a_public_workbook_load(tmp_path) -> None:
    import openpyxl
    from openpyxl.workbook.defined_name import DefinedName

    path = tmp_path / "named.xlsx"
    source = openpyxl.Workbook()
    sheet = source.active
    sheet.title = "Sheet1"
    sheet["A1"] = 41
    sheet["B1"] = "=MyValue+1"
    source.defined_names.add(DefinedName("MyValue", attr_text="'Sheet1'!$A$1"))
    source.save(path)

    wb = fz.Workbook.load_path(str(path))
    wb.evaluate_all()
    report = wb.precedents("Sheet1!B1")
    assert len(report.precedents) == 1
    reference = report.precedents[0].reference
    assert reference.kind is fz.ReferenceKind.Name
    assert reference.name == "MyValue"
    assert reference.resolution.kind is fz.NameResolutionKind.Cell
    assert reference.resolution.address == "Sheet1!A1"


def test_new_enum_member_sets_are_exact() -> None:
    expected = {
        fz.Staleness: {"Current", "Dirty", "NeverEvaluated", "Unknown"},
        fz.Provenance: {"Declared", "Observed", "Unknown"},
        fz.LinkDisposition: {"Expanded", "Convergent", "Cycle", "Elided", "Unknown"},
        fz.TraceDirection: {"Precedents", "Dependents"},
        fz.OmittedCountKind: {"Exact", "AtLeast", "Unknown"},
        fz.SpillRoleKind: {"Anchor", "Member", "Unknown"},
        fz.ReferenceKind: {
            "Cell",
            "Range",
            "Name",
            "Table",
            "External",
            "Unsupported",
            "Unknown",
        },
        fz.NameResolutionKind: {
            "Cell",
            "Range",
            "Literal",
            "Formula",
            "Unresolved",
            "Unknown",
        },
        fz.TraceLinkKindType: {"Formula", "SpillAnchor", "SpillReader", "Unknown"},
    }
    for enum_type, members in expected.items():
        runtime_members = {
            name
            for name, value in vars(enum_type).items()
            if isinstance(value, enum_type)
        }
        assert runtime_members == members
