from pathlib import Path

import openpyxl
import pytest

import formualizer as fz


def _build_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = 21
    sheet["A2"] = 5
    sheet["B1"] = "=A1*2"
    sheet["B2"] = "=A1+A2"
    workbook.save(path)


def _assert_loaded(workbook: fz.Workbook) -> None:
    assert workbook.sheet_names == ["Data"]
    sheet = workbook.sheet("Data")
    assert sheet.get_cell(1, 1).value == 21.0
    assert sheet.get_cell(2, 1).value == 5.0
    assert workbook.evaluate_cell("Data", 1, 2) == 42.0
    assert workbook.evaluate_cell("Data", 2, 2) == 26.0


def test_runtime_path_sources_match_across_public_loaders(tmp_path: Path) -> None:
    path = tmp_path / "runtime-source.xlsx"
    _build_workbook(path)

    workbooks = [
        fz.load_workbook(str(path)),
        fz.load_workbook(str(path), path_source=fz.XlsxPathSource.SHARED_FILE),
        fz.load_workbook(str(path), path_source=fz.XlsxPathSource.DIRECT_MMAP),
        fz.Workbook.load_path(str(path), path_source=fz.XlsxPathSource.SHARED_FILE),
        fz.Workbook.load_path(str(path), path_source=fz.XlsxPathSource.DIRECT_MMAP),
    ]
    for workbook in workbooks:
        _assert_loaded(workbook)


def test_path_source_rejects_invalid_enum_value(tmp_path: Path) -> None:
    path = tmp_path / "runtime-source.xlsx"
    _build_workbook(path)

    with pytest.raises(TypeError, match="XlsxPathSource"):
        fz.load_workbook(str(path), path_source="direct_mmap")


def test_direct_mmap_rejects_non_calamine_backend(tmp_path: Path) -> None:
    path = tmp_path / "runtime-source.xlsx"
    _build_workbook(path)

    with pytest.raises(ValueError, match="requires backend='calamine'"):
        fz.Workbook.load_path(
            str(path),
            backend="umya",
            path_source=fz.XlsxPathSource.DIRECT_MMAP,
        )


def test_direct_mmap_rejects_non_xlsx_path(tmp_path: Path) -> None:
    path = tmp_path / "workbook.csv"
    path.write_text("value\n42\n", encoding="utf-8")

    with pytest.raises(ValueError, match="supports only Calamine `.xlsx`"):
        fz.load_workbook(str(path), path_source=fz.XlsxPathSource.DIRECT_MMAP)
