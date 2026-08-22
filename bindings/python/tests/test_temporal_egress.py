import datetime

import pytest

import formualizer as fz

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None


def test_computed_date_native_by_default_and_serial_opt_out():
    wb = fz.Workbook()
    wb.set_formula("Sheet1", 1, 1, "=DATE(2024,12,1)")
    wb.evaluate_all()

    assert wb.get_value("Sheet1", 1, 1) == datetime.date(2024, 12, 1)

    wb.set_temporal_egress("serial")
    value = wb.get_value("Sheet1", 1, 1)
    assert isinstance(value, float)
    assert value == 45627.0


@pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")
def test_loaded_xlsx_materializes_off_origin_formulas_and_ranges(tmp_path):
    path = tmp_path / "temporal-egress.xlsx"
    source = openpyxl.Workbook()
    sheet = source.active
    sheet.title = "Sheet1"
    sheet["F10"] = datetime.date(2024, 10, 18)
    sheet["F10"].number_format = "m/d/yyyy"
    sheet["G10"] = "=F10+1"
    sheet["H11"] = "=DATE(2024,12,1)"
    sheet["I12"] = "=1+1"
    source.save(path)

    workbook = fz.load_workbook(str(path), strategy="eager_all")
    workbook.evaluate_all()

    assert workbook.get_value("Sheet1", 10, 6) == datetime.date(2024, 10, 18)
    assert workbook.get_value("Sheet1", 10, 7) == datetime.date(2024, 10, 19)
    assert workbook.get_value("Sheet1", 11, 8) == datetime.date(2024, 12, 1)
    assert workbook.get_value("Sheet1", 12, 9) == 2.0
    assert workbook.sheet("Sheet1").get_values(
        fz.RangeAddress("Sheet1", 10, 6, 10, 7)
    ) == [[datetime.date(2024, 10, 18), datetime.date(2024, 10, 19)]]
