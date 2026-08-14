"""Custom functions must work on workbooks loaded from XLSX without hanging.

Regression test for the GIL deadlock where `evaluate_all`/`evaluate_cell`/
`evaluate_cells` on a loaded workbook would block forever as soon as a rayon
worker evaluated a Python-backed custom function: the main thread held the GIL
while waiting on the parallel layer, and the worker blocked on
``PyGILState_Ensure`` waiting for that same GIL.

The workbook is sized so a single evaluation layer contains many formula
vertices, which forces the engine's parallel layer to run on rayon worker
threads and reproduces the deadlock on the broken binding.

These tests are timeout-bounded. A deadlock cannot be interrupted from Python
(the GIL holder is blocked in native code, so neither a signal handler nor a
pure-Python watchdog ever runs), which is why the suite configures
pytest-timeout with `timeout_method = "thread"`: on expiry it dumps every
thread's stack and kills the process, turning a would-be six-hour CI hang into
a bounded failure with the deadlock cycle printed.
"""

import datetime
from pathlib import Path

import pytest

import formualizer as fz

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - allow skipping if not present in dev env
    openpyxl = None

pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")

_SER0 = datetime.date(1899, 12, 30)

# Matches the benchmark XNPV reference case (LibreOffice cross-check):
# rate 0.1, cashflows -1000/200/300/400/500 on quarterly dates.
_NPV_VALUE = 308.1871372025822

_NUM_FORMULAS = 120


def _xnpv(rate, values, dates):
    def column(m):
        return [row[0] if isinstance(row, (list, tuple)) else row for row in m]

    def to_serial(x):
        if isinstance(x, datetime.date):
            return (x - _SER0).days
        return float(x)

    v = [float(x) for x in column(values)]
    d = [to_serial(x) for x in column(dates)]
    d0 = d[0]
    return sum(v[i] / (1.0 + rate) ** ((d[i] - d0) / 365.0) for i in range(len(v)))


def make_loaded_workbook(tmp: Path) -> fz.Workbook:
    path = tmp / "custom_fn.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 0.1
    values = [-1000, 200, 300, 400, 500]
    dates = [
        datetime.date(2024, 1, 1),
        datetime.date(2024, 4, 1),
        datetime.date(2024, 7, 1),
        datetime.date(2024, 10, 1),
        datetime.date(2025, 1, 1),
    ]
    for i, (value, date) in enumerate(zip(values, dates, strict=True)):
        ws.cell(row=2 + i, column=2, value=value)
        cell = ws.cell(row=2 + i, column=3, value=date)
        cell.number_format = "yyyy-mm-dd"
    for i in range(_NUM_FORMULAS):
        ws.cell(
            row=2 + i,
            column=4,
            value="=XNPV($A$1,$B$2:$B$6,$C$2:$C$6)",
        )
    wb.save(path)
    return fz.load_workbook(str(path), strategy="eager_all")


def register_xnpv(wb: fz.Workbook) -> None:
    wb.register_function(
        "xnpv", _xnpv, min_args=3, max_args=3, allow_override_builtin=True
    )


@pytest.mark.timeout(60)
def test_custom_function_on_loaded_workbook_evaluate_all(tmp_path: Path):
    wb = make_loaded_workbook(tmp_path)
    register_xnpv(wb)

    wb.evaluate_all()
    for row in (3, 42, 121):
        assert wb.get_value("Sheet1", row, 4) == pytest.approx(_NPV_VALUE, rel=1e-9)


@pytest.mark.timeout(60)
def test_custom_function_on_loaded_workbook_evaluate_cell(tmp_path: Path):
    wb = make_loaded_workbook(tmp_path)
    register_xnpv(wb)

    value = wb.evaluate_cell("Sheet1", 3, 4)
    assert value == pytest.approx(_NPV_VALUE, rel=1e-9)


@pytest.mark.timeout(60)
def test_custom_function_on_loaded_workbook_evaluate_cells(tmp_path: Path):
    wb = make_loaded_workbook(tmp_path)
    register_xnpv(wb)

    results = wb.evaluate_cells([("Sheet1", 3, 4), ("Sheet1", 121, 4)])
    assert results[0] == pytest.approx(_NPV_VALUE, rel=1e-9)
    assert results[1] == pytest.approx(_NPV_VALUE, rel=1e-9)


@pytest.mark.timeout(60)
def test_native_xnpv_handles_date_cells_without_the_override(tmp_path: Path):
    """The native XNPV must agree with the Python override on this fixture.

    This started life as ``test_native_xnpv_still_errors_without_override``,
    asserting ``#NUM!`` - true when #327 was written, because the native XNPV
    dropped the ``Date`` cells in ``C2:C6`` and hit its length guard. #328
    landed the date-serial coercion right afterwards, so on plain ``main``
    (baf5b363) that assertion is already failing::

        assert isinstance(value, dict)
        E   assert False
        E    +  where False = isinstance(308.1871372025822, dict)

    The guard the original test wanted - "the override is not masking a change
    in the native function" - is preserved by pinning the native result to the
    same oracle constant the override produces.
    """
    wb = make_loaded_workbook(tmp_path)

    wb.evaluate_all()
    value = wb.get_value("Sheet1", 3, 4)
    assert value == pytest.approx(_NPV_VALUE, rel=1e-9)
